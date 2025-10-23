from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, make_response
from werkzeug.security import check_password_hash, generate_password_hash
from app.models import Show, Booking, Settings, Ticket, Buyer, AuditLog
from app import db
from app.utils.email import send_payment_confirmed
from app.utils.tickets import generate_tickets_for_booking, delete_ticket, mark_ticket_as_used
from app.utils.audit import log_payment_confirmed, log_settings_changed
from datetime import datetime
import io
import csv
from openpyxl import Workbook

admin_bp = Blueprint('admin', __name__)

def get_concert_name():
    """Helper function to get concert name from settings"""
    return Settings.get_value('concert_name', 'Klasskonsert 24C')

def login_required(f):
    """Decorator to require admin login"""
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin.login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    """Admin login page"""
    if request.method == 'POST':
        password = request.form.get('password')
        admin_password = 'klasskonsert26'  # In production, this should be hashed
        
        if password == admin_password:
            session['admin_logged_in'] = True
            flash('Inloggning lyckades!', 'success')
            return redirect(url_for('admin.dashboard'))
        else:
            flash('Felaktigt lösenord.', 'error')
    
    return render_template('admin/login.html', concert_name=get_concert_name())

@admin_bp.route('/logout')
def logout():
    """Admin logout"""
    session.pop('admin_logged_in', None)
    flash('Du har loggats ut.', 'info')
    return redirect(url_for('admin.login'))

@admin_bp.route('/')
@login_required
def dashboard():
    """Admin dashboard with all bookings"""
    # Get filter parameter
    filter_unconfirmed = request.args.get('unconfirmed', type=bool)
    
    # Build query
    query = Booking.query
    
    if filter_unconfirmed:
        # Show only unconfirmed payments (reserved status or buyer confirmed but not admin confirmed)
        query = query.filter(
            (Booking.status == 'reserved') | 
            (Booking.buyer_confirmed_payment == True) & (Booking.status != 'confirmed')
        )
    
    bookings = query.order_by(Booking.created_at.desc()).all()
    
    # Group bookings by show
    bookings_by_show = {}
    for booking in bookings:
        show_key = f"{booking.show.start_time}-{booking.show.end_time}"
        if show_key not in bookings_by_show:
            bookings_by_show[show_key] = []
        bookings_by_show[show_key].append(booking)
    
    return render_template('admin/dashboard.html', 
                         bookings_by_show=bookings_by_show,
                         filter_unconfirmed=filter_unconfirmed,
                         concert_name=get_concert_name())

@admin_bp.route('/settings')
@login_required
def settings():
    """Admin settings page"""
    settings_data = {}
    for setting in Settings.query.all():
        settings_data[setting.key] = setting.value
    
    return render_template('admin/settings.html', settings=settings_data, concert_name=get_concert_name())

@admin_bp.route('/settings/update', methods=['POST'])
@login_required
def update_settings():
    """Update settings"""
    try:
        # Update each setting
        for key in request.form:
            if key != 'csrf_token':
                Settings.set_value(key, request.form[key])
        
        flash('Inställningar sparade!', 'success')
    except Exception as e:
        flash('Ett fel uppstod vid sparande av inställningar.', 'error')
    
    return redirect(url_for('admin.settings'))

@admin_bp.route('/booking/<int:booking_id>/confirm-payment', methods=['POST'])
@login_required
def confirm_payment(booking_id):
    """Admin confirms payment"""
    booking = Booking.query.get_or_404(booking_id)
    
    if booking.status == 'confirmed':
        flash('Betalningen är redan bekräftad.', 'info')
        return redirect(url_for('admin.dashboard'))
    
    booking.status = 'confirmed'
    booking.confirmed_at = datetime.utcnow()
    
    # Update show availability
    booking.show.update_availability()
    
    db.session.commit()
    
    # Generate individual tickets
    try:
        tickets = generate_tickets_for_booking(booking)
        flash(f'Betalning bekräftad för {booking.full_name}! {len(tickets)} biljetter genererade.', 'success')
    except Exception as e:
        flash(f'Betalning bekräftad men fel vid biljettgenerering: {str(e)}', 'warning')
    
    # Log payment confirmation
    log_payment_confirmed(booking, 'admin')
    
    # Send confirmation email to buyer
    try:
        from app.utils.email import send_payment_confirmed
        email_sent = send_payment_confirmed(booking)
        if email_sent:
            flash(f'Biljetter skickade till {booking.email}!', 'success')
        else:
            flash('Biljetter genererade men e-post kunde inte skickas.', 'warning')
    except Exception as e:
        flash(f'Biljetter genererade men fel vid e-postssändning: {str(e)}', 'warning')
    
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/booking/<int:booking_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_booking(booking_id):
    """Edit booking"""
    booking = Booking.query.get_or_404(booking_id)
    
    if request.method == 'POST':
        try:
            booking.first_name = request.form.get('first_name')
            booking.last_name = request.form.get('last_name')
            booking.email = request.form.get('email')
            booking.phone = request.form.get('phone')
            
            # Note: adult_tickets and student_tickets are not updated
            # as they cannot be changed after booking creation
            
            db.session.commit()
            flash('Bokning uppdaterad!', 'success')
            return redirect(url_for('admin.dashboard'))
        except Exception as e:
            flash('Ett fel uppstod vid uppdatering.', 'error')
    
    return render_template('admin/edit_booking.html', booking=booking, concert_name=get_concert_name())

@admin_bp.route('/booking/<int:booking_id>/delete', methods=['POST'])
@login_required
def delete_booking(booking_id):
    """Delete booking with confirmation"""
    booking = Booking.query.get_or_404(booking_id)
    
    try:
        # Update show availability before deleting
        booking.show.update_availability()
        
        db.session.delete(booking)
        db.session.commit()
        flash(f'Bokning för {booking.full_name} har raderats.', 'success')
    except Exception as e:
        flash('Ett fel uppstod vid radering.', 'error')
    
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/booking/<int:booking_id>/resend-tickets', methods=['POST'])
@login_required
def resend_tickets(booking_id):
    """Resend tickets to user"""
    booking = Booking.query.get_or_404(booking_id)
    
    if booking.status != 'confirmed':
        flash('Kan endast skicka om biljetter för bekräftade bokningar.', 'error')
        return redirect(url_for('admin.dashboard'))
    
    if not booking.tickets:
        flash('Inga biljetter att skicka om.', 'error')
        return redirect(url_for('admin.dashboard'))
    
    try:
        from app.utils.email import send_payment_confirmed
        success = send_payment_confirmed(booking)
        
        if success:
            flash(f'Biljetter skickade om till {booking.full_name}!', 'success')
        else:
            flash('Ett fel uppstod vid omssändning av biljetter.', 'error')
    except Exception as e:
        flash(f'Ett fel uppstod: {str(e)}', 'error')
    
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/export/excel')
@login_required
def export_excel():
    """Export all bookings to Excel"""
    bookings = Booking.query.order_by(Booking.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Biljettbokningar"
    
    # Headers
    headers = ['ID', 'Namn', 'E-post', 'Telefon', 'Tid', 'Ordinarie biljetter', 'Studentbiljetter', 'Totalt', 'Status', 'Betalning bekräftad', 'Datum']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, booking in enumerate(bookings, 2):
        ws.cell(row=row, column=1, value=booking.id)
        ws.cell(row=row, column=2, value=booking.full_name)
        ws.cell(row=row, column=3, value=booking.email)
        ws.cell(row=row, column=4, value=booking.phone)
        ws.cell(row=row, column=5, value=f"{booking.show.start_time}-{booking.show.end_time}")
        ws.cell(row=row, column=6, value=booking.adult_tickets)
        ws.cell(row=row, column=7, value=booking.student_tickets)
        ws.cell(row=row, column=8, value=booking.total_amount)
        ws.cell(row=row, column=9, value=booking.status)
        ws.cell(row=row, column=10, value='Ja' if booking.buyer_confirmed_payment else 'Nej')
        ws.cell(row=row, column=11, value=booking.created_at.strftime('%Y-%m-%d %H:%M'))
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=biljettbokningar_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return response

@admin_bp.route('/shows')
@login_required
def manage_shows():
    """Manage shows/times"""
    shows = Show.query.all()
    return render_template('admin/shows.html', shows=shows, concert_name=get_concert_name())

@admin_bp.route('/shows/create', methods=['POST'])
@login_required
def create_show():
    """Create new show"""
    try:
        show = Show(
            date=request.form.get('date'),
            start_time=request.form.get('start_time'),
            end_time=request.form.get('end_time'),
            total_tickets=int(request.form.get('total_tickets', 100)),
            available_tickets=int(request.form.get('total_tickets', 100))
        )
        db.session.add(show)
        db.session.commit()
        flash('Ny föreställning skapad!', 'success')
    except Exception as e:
        flash('Ett fel uppstod vid skapande av föreställning.', 'error')
    
    return redirect(url_for('admin.manage_shows'))

@admin_bp.route('/shows/<int:show_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_show(show_id):
    """Edit show"""
    show = Show.query.get_or_404(show_id)
    
    if request.method == 'POST':
        try:
            # Get form data
            new_total_tickets = int(request.form.get('total_tickets'))
            new_available_tickets = int(request.form.get('available_tickets'))
            
            # Validation
            if new_total_tickets < 0:
                flash('Totalt antal biljetter kan inte vara negativt.', 'error')
                return render_template('admin/edit_show.html', show=show, concert_name=get_concert_name())
            
            if new_available_tickets < 0:
                flash('Tillgängliga biljetter kan inte vara negativa.', 'error')
                return render_template('admin/edit_show.html', show=show, concert_name=get_concert_name())
            
            if new_available_tickets > new_total_tickets:
                flash('Tillgängliga biljetter kan inte vara fler än totalt antal biljetter.', 'error')
                return render_template('admin/edit_show.html', show=show, concert_name=get_concert_name())
            
            # Calculate how many tickets are currently booked
            confirmed_bookings = [b for b in show.bookings if b.status == 'confirmed']
            total_booked = sum(b.total_tickets for b in confirmed_bookings)
            
            # Check if we're trying to set available tickets too low
            if new_available_tickets < total_booked:
                flash(f'Kan inte sätta tillgängliga biljetter till {new_available_tickets}. Det finns redan {total_booked} bekräftade biljetter.', 'error')
                return render_template('admin/edit_show.html', show=show, concert_name=get_concert_name())
            
            # Update show
            show.total_tickets = new_total_tickets
            show.available_tickets = new_available_tickets
            
            db.session.commit()
            flash('Föreställning uppdaterad!', 'success')
            return redirect(url_for('admin.manage_shows'))
            
        except ValueError:
            flash('Ogiltiga värden. Ange endast siffror.', 'error')
        except Exception as e:
            flash('Ett fel uppstod vid uppdatering.', 'error')
    
    return render_template('admin/edit_show.html', show=show, concert_name=get_concert_name())

@admin_bp.route('/shows/<int:show_id>/delete', methods=['POST'])
@login_required
def delete_show(show_id):
    """Delete show"""
    show = Show.query.get_or_404(show_id)
    
    if show.bookings:
        flash('Kan inte radera föreställning med befintliga bokningar.', 'error')
        return redirect(url_for('admin.manage_shows'))
    
    try:
        db.session.delete(show)
        db.session.commit()
        flash('Föreställning raderad!', 'success')
    except Exception as e:
        flash('Ett fel uppstod vid radering.', 'error')
    
    return redirect(url_for('admin.manage_shows'))

@admin_bp.route('/tickets')
@login_required
def tickets():
    """List all tickets with filters"""
    show_id = request.args.get('show_id', type=int)
    used_filter = request.args.get('used')
    search = request.args.get('search', '')
    booking_ref = request.args.get('booking_ref', '')
    
    query = Ticket.query
    
    if show_id:
        query = query.filter_by(show_id=show_id)
    
    if used_filter == 'used':
        query = query.filter_by(is_used=True)
    elif used_filter == 'unused':
        query = query.filter_by(is_used=False)
    
    if booking_ref:
        query = query.join(Booking).filter(Booking.booking_reference == booking_ref)
    
    if search:
        query = query.join(Booking).filter(
            Ticket.ticket_reference.contains(search) |
            Booking.booking_reference.contains(search) |
            Booking.first_name.contains(search) |
            Booking.last_name.contains(search)
        )
    
    tickets = query.order_by(Ticket.created_at.desc()).all()
    shows = Show.query.all()
    
    return render_template('admin/tickets.html', tickets=tickets, shows=shows,
                         selected_show=show_id, used_filter=used_filter, search=search, booking_ref=booking_ref, concert_name=get_concert_name())

@admin_bp.route('/ticket/<int:ticket_id>/delete', methods=['POST'])
@login_required
def delete_ticket_admin(ticket_id):
    """Delete a ticket"""
    ticket = Ticket.query.get_or_404(ticket_id)
    reason = request.form.get('reason', 'Admin deletion')
    
    try:
        delete_ticket(ticket, 'admin', reason)
        flash(f'Biljett {ticket.ticket_reference} har raderats.', 'success')
    except ValueError as e:
        flash(str(e), 'error')
    except Exception as e:
        flash(f'Ett fel uppstod vid radering: {str(e)}', 'error')
    
    return redirect(url_for('admin.tickets'))

@admin_bp.route('/check-ticket', methods=['GET', 'POST'])
@login_required
def check_ticket():
    """Ticket checker interface"""
    if request.method == 'POST':
        ticket_ref = request.form.get('ticket_reference', '').strip().upper()
        
        if not ticket_ref:
            flash('Ange en biljettreferens.', 'error')
            return render_template('admin/check_ticket.html', concert_name=get_concert_name())
        
        ticket = Ticket.query.filter_by(ticket_reference=ticket_ref).first()
        
        if not ticket:
            flash(f'Biljett {ticket_ref} hittades inte.', 'error')
            return render_template('admin/check_ticket.html', concert_name=get_concert_name())
        
        # Toggle ticket state
        try:
            from app.utils.tickets import change_ticket_state
            change_ticket_state(ticket, 'admin')
            
            if ticket.is_used:
                flash(f'Biljett {ticket_ref} markerad som använd!', 'success')
            else:
                flash(f'Biljett {ticket_ref} återställd till oanvänd!', 'success')
        except Exception as e:
            flash(f'Ett fel uppstod: {str(e)}', 'error')
        
        return render_template('admin/check_ticket.html', ticket=ticket, concert_name=get_concert_name())
    
    return render_template('admin/check_ticket.html', concert_name=get_concert_name())

@admin_bp.route('/ticket/<int:ticket_id>/toggle-state', methods=['POST'])
@login_required
def toggle_ticket_state(ticket_id):
    """Toggle ticket state (used/unused)"""
    ticket = Ticket.query.get_or_404(ticket_id)
    
    try:
        from app.utils.tickets import change_ticket_state
        change_ticket_state(ticket, 'admin')
        
        if ticket.is_used:
            flash(f'Biljett {ticket.ticket_reference} markerad som använd!', 'success')
        else:
            flash(f'Biljett {ticket.ticket_reference} återställd till oanvänd!', 'success')
    except Exception as e:
        flash(f'Ett fel uppstod: {str(e)}', 'error')
    
    return redirect(url_for('admin.tickets'))

@admin_bp.route('/booking/<int:booking_id>/resend-confirmation', methods=['POST'])
@login_required
def resend_confirmation(booking_id):
    """Resend booking confirmation email to remind user to pay"""
    booking = Booking.query.get_or_404(booking_id)

    if booking.status == 'confirmed':
        flash('Kan endast skicka om bekräftelse för obekräftade bokningar.', 'error')
        return redirect(url_for('admin.dashboard'))

    try:
        from app.utils.email import send_booking_confirmation
        success = send_booking_confirmation(booking)

        if success:
            flash(f'Bekräftelse skickad om till {booking.full_name}!', 'success')
        else:
            flash('Ett fel uppstod vid omssändning av bekräftelse.', 'error')
    except Exception as e:
        flash(f'Ett fel uppstod: {str(e)}', 'error')

    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/audit')
@login_required
def audit_log():
    """View audit log"""
    page = request.args.get('page', 1, type=int)
    action_filter = request.args.get('action')
    entity_filter = request.args.get('entity')
    user_filter = request.args.get('user')
    
    query = AuditLog.query
    
    if action_filter:
        query = query.filter_by(action_type=action_filter)
    if entity_filter:
        query = query.filter_by(entity_type=entity_filter)
    if user_filter:
        query = query.filter_by(user_identifier=user_filter)
    
    logs = query.order_by(AuditLog.timestamp.desc()).paginate(
        page=page, per_page=50, error_out=False
    )
    
    # Get unique values for filters
    actions = db.session.query(AuditLog.action_type.distinct()).all()
    entities = db.session.query(AuditLog.entity_type.distinct()).all()
    users = db.session.query(AuditLog.user_identifier.distinct()).all()
    
    return render_template('admin/audit.html', logs=logs, 
                         actions=[a[0] for a in actions],
                         entities=[e[0] for e in entities],
                         users=[u[0] for u in users],
                         current_filters={
                             'action': action_filter,
                             'entity': entity_filter,
                             'user': user_filter
                         }, concert_name=get_concert_name())
